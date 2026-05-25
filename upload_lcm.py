"""
upload_lcm.py — Read Excel LCM files and upload records to Airtable.

Usage:
    python upload_lcm.py

Requires:
    pip install openpyxl requests
    AIRTABLE_API_KEY and AIRTABLE_BASE_ID set in environment (or .env file)
"""

import os
import sys
import time

import requests

try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ── Config ─────────────────────────────────────────────────────────────────────

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv optional — env vars may already be set

API_KEY = os.getenv("AIRTABLE_API_KEY")
BASE_ID = os.getenv("AIRTABLE_BASE_ID")

if not API_KEY:
    sys.exit("ERROR: AIRTABLE_API_KEY is not set.")
if not BASE_ID:
    sys.exit("ERROR: AIRTABLE_BASE_ID is not set.")

HEADERS = {
    "Authorization": f"Bearer {API_KEY}",
    "Content-Type": "application/json",
}

META_URL    = f"https://api.airtable.com/v0/meta/bases/{BASE_ID}/tables"
RECORDS_URL = f"https://api.airtable.com/v0/{BASE_ID}"

# ── File → table name mapping ──────────────────────────────────────────────────

FILES = [
    ("JV_Agreement_LCM_CN_SG.xlsx",                  "JV Agreement Clauses (CN-SG)"),
    ("MOU_LCM_CN_SG.xlsx",                           "MOU Clauses (CN-SG)"),
    ("Exclusive_Distribution_Agreement_LCM_CN_SG.xlsx",    "Exclusive Distribution Clauses (CN-SG)"),
    ("NonExclusive_Distribution_Agreement_LCM_CN_SG.xlsx", "Non-Exclusive Distribution Clauses (CN-SG)"),
]

SHEET_NAME = "LCM Records"

# Field name → Airtable field type
FIELD_TYPES = {
    "Concept ID":            "singleLineText",
    "Label":                 "singleLineText",
    "CN Behavior":           "multilineText",
    "SG Requirement":        "multilineText",
    "Operation":             "singleLineText",
    "Template Clause":       "multilineText",
    "Statutory Reference":   "singleLineText",
    "Statutory Detail":      "multilineText",
    "Risk Level":            "singleLineText",
    "Requires Human Review": "singleLineText",
    "Notes":                 "multilineText",
    "Automation Type":       "singleLineText",
    "Document Type":         "singleLineText",
    "Clause Photo":          "singleLineText",  # attachments handled as text refs
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _encode_table(name: str) -> str:
    return name.replace(" ", "%20")


def create_table(table_name: str, columns: list[str]) -> str:
    """Create a new Airtable table. Returns the table ID."""
    # Primary field must be first; remaining fields follow
    primary = columns[0]
    other   = columns[1:]

    fields = [{"name": primary, "type": FIELD_TYPES.get(primary, "singleLineText")}]
    fields += [{"name": col, "type": FIELD_TYPES.get(col, "singleLineText")} for col in other]

    payload = {"name": table_name, "fields": fields}
    r = requests.post(META_URL, headers=HEADERS, json=payload)

    if r.status_code == 200:
        table_id = r.json()["id"]
        print(f"  ✓ Table created: '{table_name}' (id={table_id})")
        return table_id

    if r.status_code == 422 and "already exists" in r.text.lower():
        # Table exists — fetch its ID instead
        print(f"  ~ Table already exists: '{table_name}' — fetching ID...")
        return get_table_id(table_name)

    r.raise_for_status()


def get_table_id(table_name: str) -> str:
    """Fetch the ID of an existing table by name."""
    r = requests.get(META_URL, headers=HEADERS)
    r.raise_for_status()
    for tbl in r.json().get("tables", []):
        if tbl["name"] == table_name:
            return tbl["id"]
    raise ValueError(f"Table '{table_name}' not found in base after creation attempt.")


def upload_records(table_id: str, records: list[dict]) -> int:
    """Upload records in batches of 10. Returns count of records uploaded."""
    url        = f"{RECORDS_URL}/{table_id}"
    uploaded   = 0
    batch_size = 10

    for i in range(0, len(records), batch_size):
        batch = records[i : i + batch_size]
        payload = {"records": [{"fields": rec} for rec in batch]}

        r = requests.post(url, headers=HEADERS, json=payload)

        if r.status_code != 200:
            print(f"  ✗ Batch {i//batch_size + 1} failed ({r.status_code}): {r.text[:200]}")
            continue

        uploaded += len(r.json().get("records", []))
        print(f"  · Uploaded records {i+1}–{min(i+batch_size, len(records))} / {len(records)}", end="\r")
        time.sleep(0.25)  # stay well within Airtable's 5 req/s limit

    print()  # newline after \r progress
    return uploaded


def read_excel(filepath: str) -> tuple[list[str], list[dict]]:
    """Read the LCM Records sheet. Returns (columns, rows)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found in {filepath}.\n"
            f"Available sheets: {available}"
        )

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError(f"Sheet '{SHEET_NAME}' in {filepath} is empty.")

    # First row = headers
    columns = [str(c).strip() if c is not None else f"Column_{i}" for i, c in enumerate(rows[0])]

    records = []
    for row in rows[1:]:
        # Skip entirely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {}
        for col, val in zip(columns, row):
            if val is not None and str(val).strip() != "":
                record[col] = str(val).strip()
        if record:
            records.append(record)

    return columns, records


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"\nAirtable base: {BASE_ID}")
    print(f"API key prefix: {API_KEY[:12]}...\n")

    summary = []

    for filename, table_name in FILES:
        print(f"{'─'*55}")
        print(f"File:  {filename}")
        print(f"Table: {table_name}")

        if not os.path.exists(filename):
            print(f"  ✗ File not found — skipping.\n")
            summary.append((table_name, 0, "FILE NOT FOUND"))
            continue

        try:
            columns, records = read_excel(filename)
            print(f"  · Read {len(records)} records, {len(columns)} columns from '{SHEET_NAME}'")

            table_id = create_table(table_name, columns)
            uploaded = upload_records(table_id, records)

            msg = f"{uploaded} records uploaded"
            summary.append((table_name, uploaded, "OK"))
            print(f"  ✓ {msg}")

        except Exception as e:
            print(f"  ✗ Error: {e}")
            summary.append((table_name, 0, str(e)))

    print(f"\n{'═'*55}")
    print("SUMMARY")
    print(f"{'═'*55}")
    for name, count, status in summary:
        print(f"  {name}")
        print(f"    → {count} records uploaded  [{status}]")
    print()


if __name__ == "__main__":
    main()
